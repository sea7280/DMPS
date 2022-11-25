import openpyxl
from tkinter import messagebox
import os
from osgeo import gdal
import numpy as np
import sys
sys.dont_write_bytecode = True

def saveExcel(filepath, setting_detail, ndviData, fdiData):
    def write_excel(band,sheet):
        ws = sheet
        size_y = len(band)
        size_x = len(band[0])
        for i in range(size_y):
            for j in range(size_x):
                ws.cell(row=i + 1,column=j + 1).value = band[i][j]

    minX      = setting_detail[4]
    minY      = setting_detail[5]
    deltaX    = setting_detail[6]
    deltaY    = setting_detail[6]
    string    = setting_detail[11]

    bluepath        = filepath[0]
    greenpath       = filepath[1]
    redpath         = filepath[3]
    nirpath         = filepath[4]
    bluepath_aco    = filepath[6]
    greenpath_aco   = filepath[7]
    redpath_aco     = filepath[8]
    nirpath_aco     = filepath[9]
    R_RE2path_aco   = filepath[10]
    R_SWIR1path_aco = filepath[11]
    ndvi                  = ndviData
    FDI                   = fdiData


    band2_8bit_path =os.path.dirname(__file__) + "/tif_file/Band2_8bit.tif"
    band3_8bit_path =os.path.dirname(__file__) + "/tif_file/Band3_8bit.tif"
    band4_8bit_path =os.path.dirname(__file__) + "/tif_file/Band4_8bit.tif"
    band8_8bit_path =os.path.dirname(__file__) + "/tif_file/Band8_8bit.tif"

    gdal.Translate(band2_8bit_path ,bluepath    , srcWin=[minX,minY,deltaX,deltaY])
    gdal.Translate(band3_8bit_path ,greenpath   , srcWin=[minX,minY,deltaX,deltaY])
    gdal.Translate(band4_8bit_path ,redpath     , srcWin=[minX,minY,deltaX,deltaY])
    gdal.Translate(band8_8bit_path ,nirpath     , srcWin=[minX,minY,deltaX,deltaY])

    b2_image =gdal.Open(band2_8bit_path)
    b3_image =gdal.Open(band3_8bit_path)
    b4_image =gdal.Open(band4_8bit_path)
    b8_image =gdal.Open(band8_8bit_path)

    BlueBand_array      = b2_image.ReadAsArray()
    GreenBand_array     = b3_image.ReadAsArray()
    RedBand_array       = b4_image.ReadAsArray()
    NIR_Band_array      = b8_image.ReadAsArray()

    BlueBand_array     = BlueBand_array     / 10000
    GreenBand_array    = GreenBand_array    / 10000
    RedBand_array      = RedBand_array      / 10000
    NIR_Band_array     = NIR_Band_array     / 10000


#--------------------------------- acolite --------------------------------------
    cut_blue_img  =gdal.Open(bluepath_aco   )
    cut_green_img =gdal.Open(greenpath_aco  )
    cut_red_img   =gdal.Open(redpath_aco    )
    cut_NIR_img   =gdal.Open(nirpath_aco    )
    cut_RE2_img   =gdal.Open(R_RE2path_aco  )
    cut_SWIR1_img =gdal.Open(R_SWIR1path_aco)

    Blue_Band_array_aco     = cut_blue_img.ReadAsArray()
    Green_Band_array_aco    = cut_green_img.ReadAsArray()
    Red_Band_array_aco      = cut_red_img.ReadAsArray()
    NIR_Band_array_aco      = cut_NIR_img.ReadAsArray()
    R_RE2_Band_array_aco    = cut_RE2_img.ReadAsArray()
    R_SWIR1_Band_array_aco  = cut_SWIR1_img.ReadAsArray()

################################################ 切り抜き処理 ##############################################################
    if deltaX == 0:
        pass
    else:
        raw = np.arange(minX)
        row = np.arange(minY)

        Blue_Band_array_aco = np.delete(Blue_Band_array_aco,row,axis=0)
        Blue_Band_array_aco = np.delete(Blue_Band_array_aco,raw,axis=1)
        Green_Band_array_aco = np.delete(Green_Band_array_aco,row,axis=0)
        Green_Band_array_aco = np.delete(Green_Band_array_aco,raw,axis=1)
        Red_Band_array_aco = np.delete(Red_Band_array_aco,row,axis=0)
        Red_Band_array_aco = np.delete(Red_Band_array_aco,raw,axis=1)
        NIR_Band_array_aco = np.delete(NIR_Band_array_aco,row,axis=0)
        NIR_Band_array_aco = np.delete(NIR_Band_array_aco,raw,axis=1)
        R_RE2_Band_array_aco = np.delete(R_RE2_Band_array_aco,row,axis=0)
        R_RE2_Band_array_aco = np.delete(R_RE2_Band_array_aco,raw,axis=1)
        R_SWIR1_Band_array_aco = np.delete(R_SWIR1_Band_array_aco,row,axis=0)
        R_SWIR1_Band_array_aco = np.delete(R_SWIR1_Band_array_aco,raw,axis=1)

        raw_delta = np.arange(deltaX,len(NIR_Band_array_aco[0]),1)
        row_delta = np.arange(deltaY,len(NIR_Band_array_aco),1)

        Blue_Band_array_aco = np.delete(Blue_Band_array_aco,row_delta,axis=0)
        Blue_Band_array_aco = np.delete(Blue_Band_array_aco,raw_delta,axis=1)
        Green_Band_array_aco = np.delete(Green_Band_array_aco,row_delta,axis=0)
        Green_Band_array_aco = np.delete(Green_Band_array_aco,raw_delta,axis=1)
        Red_Band_array_aco = np.delete(Red_Band_array_aco,row_delta,axis=0)
        Red_Band_array_aco = np.delete(Red_Band_array_aco,raw_delta,axis=1)
        NIR_Band_array_aco = np.delete(NIR_Band_array_aco,row_delta,axis=0)
        NIR_Band_array_aco = np.delete(NIR_Band_array_aco,raw_delta,axis=1)
        R_RE2_Band_array_aco = np.delete(R_RE2_Band_array_aco,row_delta,axis=0)
        R_RE2_Band_array_aco = np.delete(R_RE2_Band_array_aco,raw_delta,axis=1)
        R_SWIR1_Band_array_aco  = np.delete(R_SWIR1_Band_array_aco ,row_delta,axis=0)
        R_SWIR1_Band_array_aco  = np.delete(R_SWIR1_Band_array_aco ,raw_delta,axis=1)

################################################ save ##########################################################

    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]
    ws.title = "Blue"
    ws = wb.create_sheet(title="Green")
    ws = wb.create_sheet(title="Red")
    ws = wb.create_sheet(title="NIR")
    ws = wb.create_sheet(title="Blue_acolite")
    ws = wb.create_sheet(title="Green_acolite")
    ws = wb.create_sheet(title="Red_acolite")
    ws = wb.create_sheet(title="NIR_acolite")
    ws = wb.create_sheet(title="RE2_acolite")
    ws = wb.create_sheet(title="SWIR_acolite")
    ws = wb.create_sheet(title="NDVI")
    ws = wb.create_sheet(title="FDI")

    write_excel(BlueBand_array        ,wb["Blue"])
    write_excel(GreenBand_array       ,wb["Green"])
    write_excel(RedBand_array         ,wb["Red"])
    write_excel(NIR_Band_array        ,wb["NIR"])
    write_excel(Blue_Band_array_aco   ,wb["Blue_acolite"])
    write_excel(Green_Band_array_aco  ,wb["Green_acolite"])
    write_excel(Red_Band_array_aco    ,wb["Red_acolite"])
    write_excel(NIR_Band_array        ,wb["NIR_acolite"])
    write_excel(NIR_Band_array_aco    ,wb["NIR_acolite"])
    write_excel(R_RE2_Band_array_aco  ,wb["RE2_acolite"])
    write_excel(R_SWIR1_Band_array_aco,wb["SWIR_acolite"])
    write_excel(ndvi                  ,wb["NDVI"])
    write_excel(FDI                   ,wb["FDI"])

    wb.save(os.getcwd() + f"/excel/excel_{string}.xlsx")
    messagebox.showinfo('Complete', 'Write to Excel completed')
