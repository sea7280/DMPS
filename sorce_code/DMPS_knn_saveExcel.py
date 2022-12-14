import openpyxl
import os
from openpyxl.styles import PatternFill
import pickle
import sys
sys.dont_write_bytecode = True

def knnSaveExcel(string):
    with open(os.path.dirname(__file__) + "/pickle/judge.pickle", mode='rb') as f:
        judegedata = pickle.load(f)
    #エクセルに判定結果を保存
    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]
    ws.title = "判定結果"
    write_excel(judegedata,wb["判定結果"])
    wb.save(os.getcwd() + f"/judge/判定結果_{string}.xlsx")



def write_excel(band,sheet):
    ws = sheet
    size_y = len(band)
    size_x = len(band[0])
    for i in range(size_y):
        for j in range(size_x):
            if band[i][j][0] == "plastic":
                ws.cell(row=i + 1,column=j + 1).fill = PatternFill(patternType='solid', fgColor='ff3333')
            elif band[i][j][0] == "water":
                ws.cell(row=i + 1,column=j + 1).fill = PatternFill(patternType='solid', fgColor='0000FF')
            elif band[i][j][0] == "wood":
                ws.cell(row=i + 1,column=j + 1).fill = PatternFill(patternType='solid', fgColor='008000')
            elif band[i][j][0] == "pumice":
                ws.cell(row=i + 1,column=j + 1).fill = PatternFill(patternType='solid', fgColor='D2B48C')
            ws.cell(row=i + 1,column=j + 1).value = band[i][j][0]

